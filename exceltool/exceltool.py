import openpyxl
import os
import re # 导入正则表达式库用于解析
from pathlib import Path # 使用 pathlib 处理路径
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill # 导入样式类
from openpyxl.worksheet.datavalidation import DataValidation # 导入数据验证类
from openpyxl.drawing.image import Image # 导入 Image 类用于插入图片
import logging # 导入日志库

# 初始化日志
logger = logging.getLogger(__name__)
logger.setLevel(logging.DEBUG)
logger.propagate = False  # 阻止传播到root logger
handler = logging.StreamHandler()
formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
handler.setFormatter(formatter)
logger.addHandler(handler)

# --- 默认样式配置 ---
DEFAULT_TARGET_FONT = Font(name='微软雅黑', size=11)
DEFAULT_TARGET_ALIGNMENT = Alignment(horizontal='center', vertical='center', wrap_text=True)
DEFAULT_TARGET_START_ROW = 5 # 目标工作表从此行开始填充
DEFAULT_SOURCE_START_ROW = 2 # 源工作表从此行开始读取（假设标题在第1行）

def parse_initiator_receiver(text: str, key: str) -> str:
    """辅助函数，用于解析 '发起者:' 或 '接收者:' 的内容"""
    if not isinstance(text, str):
        return "" # 如果输入不是字符串则返回空
    # 使用正则表达式查找关键字及其后的冒号，并捕获剩余部分
    # 处理冒号后可能存在的空格
    match = re.search(rf"{key}\s*[:：]\s*(.*)", text, re.IGNORECASE)
    return match.group(1).strip() if match else ""

def map_source_col_to_target_col_for_merge(source_col_index):
    """将源列索引映射到目标列索引，专用于合并逻辑。"""
    # 基于新的数据映射规则：
    # 源列1（需求名称、发起者、接收者）-> 目标列1, 2, 3, 5。基于目标列5合并？
    if source_col_index == 1:
        return 5 # 基于源列1的合并来合并目标列5
    # 源列2 -> 目标列4
    elif source_col_index == 2:
        return 4 # 基于源列2的合并来合并目标列4
    # 源列3 -> 目标列6
    elif source_col_index == 3:
        return 6 # 基于源列3的合并来合并目标列6
    # 源列4 -> 目标列7
    elif source_col_index == 4:
        return 7 # 基于源列4的合并来合并目标列7
    # 源列5 -> 目标列8
    elif source_col_index == 5:
        return 8 # 基于源列5的合并来合并目标列8
    # 源列6 -> 目标列9
    elif source_col_index == 6:
        return 9 # 基于源列6的合并来合并目标列9
    # 源列7 -> 目标列10
    elif source_col_index == 7:
        return 10 # 基于源列7的合并来合并目标列10
    # 源列8 -> 目标列11
    elif source_col_index == 8:
        return 11 # 基于源列8的合并来合并目标列11
    else:
        # 源列8之后的数据不映射用于传输
        return None

def process_excel_files(
    source_excel_path: Path,
    template_excel_path: Path,
    output_excel_path: Path,
    requirement_file_name: str, # 原始需求文档名称（例如："需求规格说明书_....doc"）
    targets: str,
    necessity: str,
    source_sheet_name: str = 'Sheet1', # 默认源工作表名称
    target_sheet_name: str = '2、功能点拆分表', # 默认目标工作表名称
    target_start_row: int = DEFAULT_TARGET_START_ROW,
    source_start_row: int = DEFAULT_SOURCE_START_ROW,
    target_font: Font = DEFAULT_TARGET_FONT,
    target_alignment: Alignment = DEFAULT_TARGET_ALIGNMENT,
    architecture_diagram_path: Path = None # 新增：架构图图片路径
):
    """
    从 source_excel_path 读取数据，根据定义的规则填充到 template_excel_path 的指定工作表中，
    应用格式，复制合并单元格，将建设目标/必要性添加到另一个工作表，插入架构图（如果提供），并保存到 output_excel_path。

    Args:
        source_excel_path (Path): 源 Excel 文件的路径 (例如：{req_base}.xlsx)。
        template_excel_path (Path): 模板 Excel 文件的路径。
        output_excel_path (Path): 最终输出 Excel 文件的保存路径。
        requirement_file_name (str): 原始需求文档文件的名称。
        targets (str): “建设目标”部分的文本。
        necessity (str): “建设必要性”部分的文本。
        source_sheet_name (str): 源文件中要读取的工作表名称。
        target_sheet_name (str): 模板文件中要写入的工作表名称。
        target_start_row (int): 在目标工作表中开始写入的行号。
        source_start_row (int): 在源工作表中开始读取的行号。
        target_font (Font): 应用于目标单元格的字体样式。
        target_alignment (Alignment): 应用于目标单元格的对齐样式。
        architecture_diagram_path (Path, optional): 架构图图片的路径。默认为 None。
    """
    logger.info(f"开始处理 Excel 文件...")
    logger.info(f"源文件: {source_excel_path}")
    logger.info(f"模板文件: {template_excel_path}")
    logger.info(f"输出文件: {output_excel_path}")

    if not source_excel_path.exists():
        logger.error(f"错误：源文件 '{source_excel_path}' 不存在。")
        return False
    if not template_excel_path.exists():
        logger.error(f"错误：模板文件 '{template_excel_path}' 不存在。")
        return False

    source_wb = None
    template_wb = None

    try:
        logger.info("正在加载工作簿...")
        source_wb = openpyxl.load_workbook(source_excel_path)
        template_wb = openpyxl.load_workbook(template_excel_path)
        logger.info("工作簿加载完成。")

        logger.info(f"正在选择工作表...")
        if source_sheet_name not in source_wb.sheetnames:
             logger.error(f"错误：源文件 '{source_excel_path}' 中找不到名为 '{source_sheet_name}' 的工作表。")
             return False
        source_ws = source_wb[source_sheet_name]

        try:
            target_ws = template_wb[target_sheet_name]
        except KeyError:
            sheet_names = template_wb.sheetnames
            if len(sheet_names) >= 3:
                target_sheet_index = 2 # Index 2 corresponds to the 3rd sheet
                actual_target_sheet_name = sheet_names[target_sheet_index]
                logger.warning(f"警告：在模板文件 '{template_excel_path}' 中找不到名为 '{target_sheet_name}' 的工作表。")
                logger.warning(f"将使用第三个工作表：'{actual_target_sheet_name}' (索引 {target_sheet_index})。")
                target_ws = template_wb.worksheets[target_sheet_index]
            else:
                logger.error(f"错误：模板文件 '{template_excel_path}' 中找不到名为 '{target_sheet_name}' 的工作表，且工作表总数少于3个。")
                return False

        logger.info(f"源工作表: '{source_ws.title}'")
        logger.info(f"目标工作表: '{target_ws.title}'")

        logger.info(f"开始从源文件第 {source_start_row} 行读取数据，填充到目标文件第 {target_start_row} 行...")
        current_target_row = target_start_row
        processed_rows_count = 0
        max_source_row = source_ws.max_row

        # 定义要格式化的目标列（1到13）
        target_columns_to_format = list(range(1, 14))

        # 为目标列1准备需求名称
        req_name_cleaned = requirement_file_name.replace("需求规格说明书_", "").split('.')[0] # 移除前缀和扩展名

        for s_row_index in range(source_start_row, max_source_row + 1):
            # --- 1. 数据提取与解析 ---
            source_col1_val = source_ws.cell(row=s_row_index, column=1).value
            source_col2_val = source_ws.cell(row=s_row_index, column=2).value
            source_col3_val = source_ws.cell(row=s_row_index, column=3).value
            # 读取源列3到8用于目标列6到11
            source_cols_3_to_8 = [source_ws.cell(row=s_row_index, column=c).value for c in range(3, 9)] # C列到H列

            initiator = parse_initiator_receiver(source_col1_val, "发起者")
            receiver = parse_initiator_receiver(source_col1_val, "接收者")

            # --- 2. 数据填充 ---
            # 目标列1：清理后的需求名称
            target_ws.cell(row=current_target_row, column=1, value=req_name_cleaned)
            # 目标列2：发起者
            target_ws.cell(row=current_target_row, column=2, value=initiator)
            # 目标列3：接收者
            target_ws.cell(row=current_target_row, column=3, value=receiver)
            # 目标列4：源列2
            target_ws.cell(row=current_target_row, column=4, value=source_col2_val)
            # 目标列5：源列1（原始值）
            target_ws.cell(row=current_target_row, column=5, value=source_col1_val)
            # 目标列6-11：源列3-8
            for i, val in enumerate(source_cols_3_to_8):
                 target_ws.cell(row=current_target_row, column=6 + i, value=val) # 6+0=6, 6+1=7,... 6+4=10


            # 目标列12：固定值 "新增"
            target_ws.cell(row=current_target_row, column=12, value="新增")
            # 目标列13：固定值 "1"
            target_ws.cell(row=current_target_row, column=13, value="1") # 存储为字符串还是数字？为保持一致性使用字符串

            # --- 3. 应用格式化 ---
            for col_idx in target_columns_to_format:
                try:
                    cell = target_ws.cell(row=current_target_row, column=col_idx)
                    cell.font = target_font
                    cell.alignment = target_alignment
                except Exception as format_error:
                    logger.warning(f"警告：无法应用格式到单元格 {get_column_letter(col_idx)}{current_target_row}。错误: {format_error}")

            current_target_row += 1
            processed_rows_count += 1

        logger.info(f"数据填充和格式应用完成，共处理了 {processed_rows_count} 行数据。")

        # --- 4a. 应用特定目标合并 ---
        last_filled_row = current_target_row - 1
        if processed_rows_count > 1 and last_filled_row >= target_start_row:
            # 合并目标列 A (列 1)
            try:
                col_a_range = f"A{target_start_row}:A{last_filled_row}"
                logger.info(f"  应用特定合并: 目标列 A ({col_a_range})")
                target_ws.merge_cells(start_row=target_start_row, start_column=1, end_row=last_filled_row, end_column=1)
                # 将格式应用于合并区域的左上角单元格
                top_left_cell_a = target_ws.cell(row=target_start_row, column=1)
                top_left_cell_a.font = target_font
                top_left_cell_a.alignment = target_alignment
            except Exception as merge_a_error:
                logger.warning(f"  警告：无法合并目标列 A ({col_a_range})。错误：{merge_a_error}")

        # --- 4b. 处理合并单元格（从源复制并扩展）---
        logger.info("开始处理从源文件复制并扩展的合并单元格...")
        merged_cells_copied = 0
        # 确保获取列表副本，因为在迭代时修改合并可能导致问题
        source_merged_ranges = list(source_ws.merged_cells.ranges)

        for merged_range in source_merged_ranges:
            # 检查合并范围是否在我们处理的行内开始
            if merged_range.min_row >= source_start_row:
                # 计算对应的目标行范围
                target_min_r = merged_range.min_row - source_start_row + target_start_row
                target_max_r = merged_range.max_row - source_start_row + target_start_row

                # 根据映射确定目标列范围
                target_min_c = float('inf')
                target_max_c = float('-inf')
                valid_map_found = False

                for s_col in range(merged_range.min_col, merged_range.max_col + 1):
                    t_col = map_source_col_to_target_col_for_merge(s_col)
                    if t_col is not None:
                        target_min_c = min(target_min_c, t_col)
                        target_max_c = max(target_max_c, t_col)
                        valid_map_found = True

                # 如果找到有效映射且不是单个单元格，则在目标工作表中应用合并
                if valid_map_found and target_min_c <= target_max_c and target_min_r <= target_max_r:
                    if not (target_min_r == target_max_r and target_min_c == target_max_c):
                        try:
                            target_range_str = f"{get_column_letter(target_min_c)}{target_min_r}:{get_column_letter(target_max_c)}{target_max_r}"
                            logger.info(f"  应用合并: 源 {merged_range.coord} -> 目标 {target_range_str}")
                            # 重要：如果目标范围与现有合并重叠，请先取消合并
                            # 这很复杂；更简单的方法是直接尝试合并。openpyxl可以处理一些重叠。
                            # 为了稳健处理，需要检查与 target_ws.merged_cells 的重叠
                            target_ws.merge_cells(start_row=target_min_r, start_column=target_min_c,
                                                  end_row=target_max_r, end_column=target_max_c)
                            merged_cells_copied += 1
                            # 将格式应用于主要合并范围的左上角单元格
                            top_left_cell = target_ws.cell(row=target_min_r, column=target_min_c)
                            top_left_cell.font = target_font
                            top_left_cell.alignment = target_alignment

                            # --- 基于列E的合并，应用列B和C的条件合并 ---
                            # 检查主要合并是否应用于列E（target_min_c == 5 且 target_max_c == 5）
                            if target_min_c == 5 and target_max_c == 5:
                                # 合并列B（列2）与相同的行范围
                                try:
                                    col_b_range = f"B{target_min_r}:B{target_max_r}"
                                    logger.info(f"    扩展合并: 目标列 B ({col_b_range}) 基于 E 列合并")
                                    target_ws.merge_cells(start_row=target_min_r, start_column=2, end_row=target_max_r, end_column=2)
                                    top_left_cell_b = target_ws.cell(row=target_min_r, column=2)
                                    top_left_cell_b.font = target_font
                                    top_left_cell_b.alignment = target_alignment
                                except Exception as merge_b_error:
                                    logger.warning(f"    警告：无法合并目标列 B ({col_b_range})。错误：{merge_b_error}")

                                # 合并列C（列3）与相同的行范围
                                try:
                                    col_c_range = f"C{target_min_r}:C{target_max_r}"
                                    logger.info(f"    扩展合并: 目标列 C ({col_c_range}) 基于 E 列合并")
                                    target_ws.merge_cells(start_row=target_min_r, start_column=3, end_row=target_max_r, end_column=3)
                                    top_left_cell_c = target_ws.cell(row=target_min_r, column=3)
                                    top_left_cell_c.font = target_font
                                    top_left_cell_c.alignment = target_alignment
                                except Exception as merge_c_error:
                                    logger.warning(f"    警告：无法合并目标列 C ({col_c_range})。错误：{merge_c_error}")

                        except Exception as merge_error:
                            # 记录错误，例如，由于合并重叠
                            logger.warning(f"  警告：无法合并主要目标范围 {target_range_str}。可能与其他合并冲突。错误：{merge_error}")

        logger.info(f"合并单元格处理完成，尝试应用了 {merged_cells_copied} 个主要合并范围（并可能扩展了 B/C 列）。")

        # --- 5. 对目标工作表应用特定格式和数据验证 ---
        logger.info(f"开始应用特定格式和数据验证到工作表 '{target_ws.title}'...")
        if last_filled_row >= target_start_row: # 检查是否处理了任何数据行
            # (1 & 2) 列宽
            logger.info("  设置列宽...")
            fixed_width_cm = 4.5
            # 近似转换：Excel宽度单位大约是默认字体中一个'0'的宽度。
            # Calibri 11 的一个常见近似值是 width = cm * 3.8
            fixed_width_unit = fixed_width_cm * 3.8 # 根据需要调整乘数
            for col_idx in range(1, 8): # A 到 G
                col_letter = get_column_letter(col_idx)
                target_ws.column_dimensions[col_letter].width = fixed_width_unit
            for col_idx in range(8, 15): # H 到 N
                col_letter = get_column_letter(col_idx)
                target_ws.column_dimensions[col_letter].auto_size = True # 尝试自动调整大小

            # (3) 对齐
            logger.info("  设置单元格对齐 (A-M)...")
            center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
            for row_idx in range(target_start_row, last_filled_row + 1):
                for col_idx in range(1, 14): # A 到 M
                    target_ws.cell(row=row_idx, column=col_idx).alignment = center_align

            # (4) M列颜色填充
            logger.info("  设置M列背景颜色...")
            fill_color = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
            for row_idx in range(target_start_row, last_filled_row + 1):
                target_ws.cell(row=row_idx, column=13).fill = fill_color # M列是第13列

            # (5) L列的下拉列表
            logger.info("  设置L列下拉列表...")
            dv = DataValidation(type="list", formula1='"新增,复用,利旧"', allow_blank=True)
            # 将验证添加到工作表
            target_ws.add_data_validation(dv)
            # 将验证应用于范围 L<start_row>:L<last_row> (L is column 12)
            dv.add(f'L{target_start_row}:L{last_filled_row}')

            # (6) M列的公式 (Moved from N, references L)
            logger.info("  设置M列公式...")
            for row_idx in range(target_start_row, last_filled_row + 1):
                # 引用同一行L列（索引12）的公式
                formula = f'=IF(L{row_idx}="新增",1,IF(L{row_idx}="复用",1/3,IF(L{row_idx}="利旧",0)))'
                target_ws.cell(row=row_idx, column=13).value = formula # M列是第13列
            # Column N (14) no longer has a formula set here.
        else:
             logger.warning(f"工作表 '{target_ws.title}' 中没有处理数据行，跳过特定格式化。")

        logger.info("特定格式和数据验证应用完成。")


        # --- 6. 处理第二个工作表（建设目标和必要性）---
        # 从 5 重新编号为 6
        if len(template_wb.sheetnames) >= 2:
            # 假设第二个工作表（索引1）是目标
            env_sheet = template_wb.worksheets[1]
            env_sheet_title = env_sheet.title
            logger.info(f"正在处理第二个工作表: '{env_sheet_title}' 用于写入建设目标和必要性。")

            # 将建设目标写入A2，合并A2:J3
            env_sheet.cell(row=2, column=1, value=targets)
            env_sheet.merge_cells(start_row=2, start_column=1, end_row=3, end_column=10)
            cell_a2 = env_sheet.cell(row=2, column=1)
            cell_a2.font = target_font
            cell_a2.alignment = target_alignment
            logger.info(f"已将建设目标写入 '{env_sheet_title}'!A2 并合并 A2:J3。")

            # 将建设必要性写入A5，合并A5:J6
            env_sheet.cell(row=5, column=1, value=necessity)
            env_sheet.merge_cells(start_row=5, start_column=1, end_row=6, end_column=10)
            cell_a5 = env_sheet.cell(row=5, column=1)
            cell_a5.font = target_font
            cell_a5.alignment = target_alignment
            logger.info(f"已将建设必要性写入 '{env_sheet_title}'!A5 并合并 A5:J6。")
        else:
            logger.warning("警告：模板文件工作表少于2个，跳过建设目标和必要性写入。")

        # --- 7. 插入架构图 (如果提供了路径) ---
        # 从 6 重新编号为 7
        if architecture_diagram_path and architecture_diagram_path.exists():
            logger.info(f"正在尝试插入架构图到第二个工作表的 A40 单元格: {architecture_diagram_path}")
            try:
                # 获取第二个工作表（索引为1）
                if len(template_wb.sheetnames) >= 2:
                    diagram_ws = template_wb.worksheets[1] # Index 1 is the second sheet
                    logger.info(f"将架构图插入到第二个工作表: '{diagram_ws.title}'")
                else:
                    logger.warning("模板文件的工作表少于2个，无法将架构图插入到第二个工作表。将跳过图片插入。")
                    diagram_ws = None # Indicate that the sheet wasn't found

                if diagram_ws:
                    img = Image(architecture_diagram_path)
                    # 可以调整图片大小，例如：
                    # img.width = img.width * 0.75
                    # img.height = img.height * 0.75
                    diagram_ws.add_image(img, 'A40') # 将图片添加到 A40 单元格
                    logger.info(f"架构图已成功添加到工作表 '{diagram_ws.title}' 的 A40 单元格。")
            except Exception as img_err:
                logger.error(f"插入架构图时出错: {img_err}", exc_info=True)
        elif architecture_diagram_path:
            logger.warning(f"提供的架构图路径不存在，无法插入图片: {architecture_diagram_path}")
        else:
            logger.info("未提供架构图路径，跳过图片插入。")


        # --- 8. 保存输出 ---
        # 从 7 重新编号为 8
        logger.info(f"正在保存最终结果到 '{output_excel_path}'...")
        template_wb.save(output_excel_path)
        logger.info("文件保存成功！")
        return True # 表示成功

    except FileNotFoundError as e:
        logger.error(f"错误：文件未找到 - {e}")
        return False
    except KeyError as e:
         logger.error(f"错误：找不到工作表 - {e}。请检查工作表名称。")
         return False
    except Exception as e:
        logger.error(f"处理 Excel 文件过程中发生未知错误：{e}", exc_info=True) # 记录回溯信息
        # import traceback
        # traceback.print_exc()
        return False
    finally:
        # 确保工作簿已关闭
        if source_wb:
            try:
                source_wb.close()
            except Exception as close_err:
                 logger.warning(f"关闭源工作簿时出错: {close_err}")
        if template_wb:
            try:
                template_wb.close()
            except Exception as close_err:
                 logger.warning(f"关闭模板工作簿时出错: {close_err}")
        logger.info("工作簿关闭操作完成。")
