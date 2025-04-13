import logging
from pathlib import Path
import os
import re
import json
from typing import Optional # Import Optional for type hinting
import openpyxl # 导入 openpyxl
import shutil # 导入 shutil 用于移动文件

from project_paths import ProjectPaths
from read_file_content import merge_temp_files, save_content_to_file, read_file_content
from exceltool.exceltool import process_excel_files
from create_req_word import generate_word_document # 导入 Word 生成函数

logger = logging.getLogger(__name__)


def run_post_processing(req_name: str, config: ProjectPaths, dev_name: str, req_base: str, doc_file: Path):
    """
    处理单个需求的后置处理步骤：
    - 合并临时COSMIC表
    - 生成最终的Markdown、Excel和Word文件
    - 生成最终的Markdown、Excel文件
    - 填充Excel模板文件
    - 生成最终的Word文件 (新)
    - 清理临时文件
    """
    logger.info(f"[{req_name}] 阶段 4: 开始后置处理")
    try:
        # Determine paths (保持不变)
        output_dir_for_req = config.output / dev_name / req_base
        output_dir_for_req.mkdir(parents=True, exist_ok=True)
        final_output_file = output_dir_for_req / f"{req_base}_cosmic_merged.md"
        temp_file_pattern = "table_cosmic_*.md"

        merged_content_for_export = None

        # Check if merged file exists
        if final_output_file.exists():
            logger.info(f"[{req_name}] 发现已存在的合并文件: {final_output_file}，将直接使用其内容。")
            try:
                merged_content_for_export = read_file_content(final_output_file)
                # Clean up lingering temp files
                lingering_temps = list(output_dir_for_req.glob(temp_file_pattern))
                if lingering_temps:
                    logger.info(f"[{req_name}] 清理残留的临时文件...")
                    deleted_count = 0
                    for temp_file in lingering_temps:
                        try:
                            temp_file.unlink()
                            deleted_count += 1
                        except OSError as e:
                            logger.warning(f"[{req_name}] 删除残留临时文件失败 {temp_file}: {e}")
                    if deleted_count > 0:
                        logger.info(f"[{req_name}] 已清理 {deleted_count} 个残留临时文件。")
            except Exception as read_err:
                logger.error(f"[{req_name}] 读取已存在的合并文件失败: {read_err}，将尝试重新合并。")
                merged_content_for_export = None

        # Merge if needed
        if merged_content_for_export is None:
            logger.info(f"[{req_name}] 未找到或无法读取合并文件，开始查找并合并临时文件...")
            temp_files_to_merge = sorted(list(output_dir_for_req.glob(temp_file_pattern)))

            if not temp_files_to_merge:
                logger.warning(f"[{req_name}] 在目录 {output_dir_for_req} 中未找到要合并的临时文件 (匹配模式: {temp_file_pattern})。无法生成合并文件和后续导出。")
            else:
                logger.info(f"[{req_name}] 找到 {len(temp_files_to_merge)} 个临时文件准备合并。")
                try:
                    merged_content_for_export = merge_temp_files(temp_files_to_merge)
                    with open(final_output_file, "w", encoding="utf-8") as f_out:
                        f_out.write(merged_content_for_export)
                    logger.info(f"[{req_name}] COSMIC 表已合并到: {final_output_file}")

                    # Delete temp files
                    logger.info(f"[{req_name}] 开始删除临时 COSMIC 表文件...")
                    deleted_count = 0
                    for temp_file in temp_files_to_merge:
                        try:
                            temp_file.unlink()
                            deleted_count += 1
                        except OSError as e:
                            logger.warning(f"[{req_name}] 删除临时文件失败 {temp_file}: {e}")
                    logger.info(f"[{req_name}] 已成功删除 {deleted_count} 个临时文件。")
                except Exception as merge_err:
                     logger.error(f"[{req_name}] 合并或写入文件过程中出错: {merge_err}")
                     merged_content_for_export = None

        # Export if content is available (保持不变)
        if merged_content_for_export is not None:
            # 5. Generate Excel (保持不变)
            logger.info(f"[{req_name}] 阶段 5: 开始将合并后的 Markdown 转换为 Excel 文件...")
            excel_output_path = output_dir_for_req / (req_base + '.xlsx')
            save_content_to_file(
                file_name=req_base,
                output_dir=output_dir_for_req,
                content=merged_content_for_export,
                content_type="xlsx"
            )
            logger.info(f"[{req_name}] Excel 文件已生成: {excel_output_path}")

            # 6. Fill template Excel (原步骤7，提前以便获取架构图路径等信息)
            logger.info(f"[{req_name}] 阶段 6: 开始使用生成的 Excel 填充模板...")
            source_excel_for_template = excel_output_path # 使用上面生成的Excel路径
            template_excel = config.base_dir / "out_template" / "template.xlsx"
            # template_word 路径将在生成Word时定义
            final_excel_output = output_dir_for_req / f"{req_base}_COSMIC.xlsx"

            # Read targets and necessity from JSON file (保持不变)
            extracted_targets: str = ""
            extracted_necessity: str = ""
            # 定义需求描述 JSON 文件路径 (与 requirement_analysis.py 中的 get_description_json_path 逻辑一致)
            json_file_path = output_dir_for_req / f"{config.REQ_DESC_PREFIX}{req_base}{config.REQ_DESC_SUFFIX}"
            logger.info(f"[{req_name}] 正在读取 JSON 文件以提取建设目标和必要性: {json_file_path}")
            if json_file_path.exists():
                try:
                    with open(json_file_path, 'r', encoding='utf-8') as f:
                        data = json.load(f)
                    if 'requirement_description' in data:
                        if 'construction_goals' in data['requirement_description'] and isinstance(data['requirement_description']['construction_goals'], list):
                            extracted_targets = "\n".join(data['requirement_description']['construction_goals'])
                            logger.info(f"[{req_name}] 已提取建设目标。")
                        else:
                            logger.warning(f"[{req_name}] 在 JSON 文件中未找到 'construction_goals' 列表。")

                        if 'necessity' in data['requirement_description'] and isinstance(data['requirement_description']['necessity'], list):
                            extracted_necessity = "\n".join(data['requirement_description']['necessity'])
                            logger.info(f"[{req_name}] 已提取建设必要性。")
                        else:
                            logger.warning(f"[{req_name}] 在 JSON 文件中未找到 'necessity' 列表。")
                    else:
                        logger.warning(f"[{req_name}] 在 JSON 文件中未找到 'requirement_description' 对象。")

                except json.JSONDecodeError:
                    logger.error(f"[{req_name}] 解析 JSON 文件失败: {json_file_path}")
                except Exception as e:
                    logger.error(f"[{req_name}] 读取或处理 JSON 文件时出错 {json_file_path}: {e}", exc_info=True)
            else:
                logger.warning(f"[{req_name}] 需求描述 JSON 文件未找到，无法提取建设目标和必要性: {json_file_path}")


            # Find architecture diagram image (保持不变, 但移到填充Excel之前)
            architecture_diagram_file: Optional[Path] = None
            # 定义架构图文件路径 (与 requirement_analysis.py 中的 get_architecture_diagram_path 逻辑一致)
            expected_diagram_path = output_dir_for_req / f"{req_base}{config.ARCH_DIAGRAM_SUFFIX}"
            logger.info(f"[{req_name}] 正在查找架构图文件: {expected_diagram_path}")
            if expected_diagram_path.exists():
                architecture_diagram_file = expected_diagram_path
                logger.info(f"[{req_name}] 找到架构图文件: {architecture_diagram_file}")
            else:
                logger.info(f"[{req_name}] 在 {output_dir_for_req} 未找到架构图文件: {expected_diagram_path.name}")


            # 填充Excel模板 (保持不变)
            if source_excel_for_template.exists() and template_excel.exists():
                success_excel_fill = process_excel_files(
                    source_excel_path=source_excel_for_template,
                    template_excel_path=template_excel,
                    output_excel_path=final_excel_output,
                    requirement_file_name=doc_file.name, # 原始需求文件名
                    targets=extracted_targets,
                    necessity=extracted_necessity,
                    architecture_diagram_path=architecture_diagram_file # Pass the image path
                )
                if success_excel_fill:
                    logger.info(f"[{req_name}] 最终模板填充 Excel 文件（包含架构图，如果找到）已生成: {final_excel_output}")
                else:
                    logger.error(f"[{req_name}] 填充模板 Excel 文件失败。")
            else:
                if not source_excel_for_template.exists():
                    logger.error(f"[{req_name}] 无法填充模板，因为源 Excel 文件不存在: {source_excel_for_template}")
                if not template_excel.exists():
                     logger.error(f"[{req_name}] 无法填充模板，因为模板 Excel 文件不存在: {template_excel}")

            # 7. Generate final Word document
            logger.info(f"[{req_name}] 阶段 7: 开始生成最终 Word 文档...")
            # 定义最终 Word 输出路径 (与 requirement_analysis.py 中的 get_final_word_path 逻辑一致)
            final_word_output = output_dir_for_req / f"{req_base}{config.FINAL_WORD_SUFFIX}"
            template_word = config.base_dir / "out_template" / "template.docx" # Word 模板路径

            # 检查依赖项
            if not json_file_path.exists():
                logger.error(f"[{req_name}] 无法生成 Word 文档，依赖的需求描述 JSON 文件不存在: {json_file_path}")
            elif not template_word.exists():
                logger.error(f"[{req_name}] 无法生成 Word 文档，模板文件不存在: {template_word}")
            else:
                # 架构图路径已在上面查找过 (architecture_diagram_file)
                img_path_for_word = str(architecture_diagram_file) if architecture_diagram_file else None
                if not img_path_for_word:
                     logger.warning(f"[{req_name}] 未找到架构图文件，Word 文档将不包含图片。")

                # 1. 处理需求标题 (移除前缀)
                requirement_title = req_base
                prefix_to_remove = "需求规格说明书_"
                if requirement_title.startswith(prefix_to_remove):
                    requirement_title = requirement_title[len(prefix_to_remove):]
                logger.info(f"[{req_name}] 使用处理后的需求标题: {requirement_title}")

                # 2. 从 *_COSMIC.xlsx 文件读取并格式化功能点文本 (替换原有逻辑)
                functional_points_text = ""
                cosmic_excel_path = final_excel_output # 使用之前定义的填充模板后的 Excel 文件路径
                logger.info(f"[{req_name}] 尝试从 Excel 文件读取功能点: {cosmic_excel_path}")

                if cosmic_excel_path.exists():
                    try:
                        workbook = openpyxl.load_workbook(cosmic_excel_path)
                        # 假设数据在第三个工作表
                        if len(workbook.sheetnames) >= 3:
                            sheet = workbook[workbook.sheetnames[2]] # 获取第三个工作表
                            logger.info(f"[{req_name}] 读取工作表: {sheet.title}")

                            formatted_lines = []
                            current_l1_title = ""
                            current_l2_title = ""
                            l1_count = 0
                            l2_count = 0
                            l3_count = 0
                            last_l1_val = ""
                            last_l2_val = ""

                            # 迭代行，从第五行开始读取 (跳过前4行标题)
                            for row_idx in range(5, sheet.max_row + 1):
                                cell_f = sheet.cell(row=row_idx, column=6) # F列
                                cell_g = sheet.cell(row=row_idx, column=7) # G列
                                cell_h = sheet.cell(row=row_idx, column=8) # H列

                                val_f = cell_f.value if cell_f.value else ""
                                val_g = cell_g.value if cell_g.value else ""
                                val_h = cell_h.value if cell_h.value else ""

                                # 处理合并单元格：如果当前单元格为空，则使用上一行的有效值
                                # 注意：openpyxl 不直接提供合并单元格信息，这种方法是基于内容的近似处理
                                if not val_f and last_l1_val: val_f = last_l1_val
                                if not val_g and last_l2_val: val_g = last_l2_val

                                # 更新当前层级标题和计数器
                                if val_f and val_f != current_l1_title:
                                    current_l1_title = val_f
                                    l1_count += 1
                                    l2_count = 0 # 重置下级计数器
                                    formatted_lines.append(f"{l1_count} {current_l1_title}")
                                    last_l1_val = val_f # 记录最后有效的L1值
                                    last_l2_val = "" # L1变了，L2也重置

                                if val_g and val_g != current_l2_title:
                                    current_l2_title = val_g
                                    l2_count += 1
                                    l3_count = 0 # 重置下级计数器
                                    formatted_lines.append(f"{l1_count}.{l2_count} {current_l2_title}")
                                    last_l2_val = val_g # 记录最后有效的L2值
                                else:
                                    # 如果 G 列值与当前 L2 标题相同或为空，但 F 列值是新的，则重置 L2 标题
                                    if val_f and val_f == current_l1_title and val_f != last_l1_val:
                                         current_l2_title = "" # 避免沿用旧的L2标题
                                         last_l2_val = ""


                                if val_h:
                                    l3_count += 1
                                    formatted_lines.append(f"{l1_count}.{l2_count}.{l3_count} {val_h}")

                                # 更新上一行的有效值记录 (即使当前行为空也更新，以便下一行判断)
                                if val_f: last_l1_val = val_f
                                if val_g: last_l2_val = val_g


                            functional_points_text = "\n".join(formatted_lines)
                            logger.info(f"[{req_name}] 成功从 Excel 格式化功能点文本。")

                        else:
                            logger.warning(f"[{req_name}] Excel 文件 '{cosmic_excel_path.name}' 的工作表数量少于3个，无法读取功能点。")
                            functional_points_text = "[无法读取功能点：工作表不足]"

                    except Exception as excel_err:
                        logger.error(f"[{req_name}] 读取或处理 Excel 文件 '{cosmic_excel_path.name}' 时出错: {excel_err}", exc_info=True)
                        functional_points_text = f"[无法读取功能点：处理 Excel 时出错 - {excel_err}]"
                else:
                    logger.warning(f"[{req_name}] 无法找到 Excel 文件 '{cosmic_excel_path.name}'，无法提取功能点。")
                    functional_points_text = "[无法读取功能点：Excel 文件未找到]"


                try:
                    # 调用 Word 生成函数 (使用从 Excel 获取的 functional_points_text)
                    success_word_gen = generate_word_document(
                        requirement_name=requirement_title, # 使用处理后的标题
                        json_data_path=str(json_file_path),
                        template_path=str(template_word),
                        word_text=functional_points_text, # 新增：传递合并后的 Markdown 内容
                        output_doc_path=str(final_word_output),
                        image_path=img_path_for_word,
                        image_placeholder='sequence_diagram_mermaid', # 与 requirement_analysis 中保持一致
                        # image_width_cm=15 # 可选参数
                    )

                    if success_word_gen:
                        logger.info(f"[{req_name}] 最终 Word 文档生成成功: {final_word_output}")
                    else:
                        logger.error(f"[{req_name}] 最终 Word 文档生成失败 (由 generate_word_document 报告)")

                except Exception as word_err:
                    logger.error(f"[{req_name}] 生成最终 Word 文档过程中发生意外错误: {word_err}", exc_info=True)

        else: # This else corresponds to the `if merged_content_for_export is not None:` block
             logger.warning(f"[{req_name}] 由于未能获取合并后的内容，跳过 Excel、Word 和最终模板填充文件生成。")

        # 8. 创建 final_result 目录并移动最终文件
        logger.info(f"[{req_name}] 阶段 8: 开始整理最终输出文件...")
        final_result_dir = output_dir_for_req / "final_result"
        try:
            final_result_dir.mkdir(parents=True, exist_ok=True)
            logger.info(f"[{req_name}] 已创建或确认目录: {final_result_dir}")

            # 移动 Excel 文件 (final_excel_output 在步骤6定义)
            excel_to_move = output_dir_for_req / f"{req_base}_COSMIC.xlsx" # 重新获取路径确保准确
            if excel_to_move.exists():
                try:
                    shutil.move(str(excel_to_move), str(final_result_dir / excel_to_move.name))
                    logger.info(f"[{req_name}] 已移动文件: {excel_to_move.name} 到 {final_result_dir.name}")
                except Exception as move_excel_err:
                    logger.error(f"[{req_name}] 移动 Excel 文件 {excel_to_move.name} 失败: {move_excel_err}", exc_info=True)
            else:
                # 检查 final_excel_output 是否存在于 locals() 以防万一
                if 'final_excel_output' in locals() and final_excel_output.exists():
                     try:
                         shutil.move(str(final_excel_output), str(final_result_dir / final_excel_output.name))
                         logger.info(f"[{req_name}] 已移动文件 (备用路径): {final_excel_output.name} 到 {final_result_dir.name}")
                     except Exception as move_excel_err_alt:
                         logger.error(f"[{req_name}] 移动 Excel 文件 (备用路径) {final_excel_output.name} 失败: {move_excel_err_alt}", exc_info=True)
                else:
                     logger.warning(f"[{req_name}] 未找到最终 Excel 文件进行移动: {excel_to_move.name}")


            # 移动 Word 文件 (final_word_output 在步骤7定义)
            word_to_move = output_dir_for_req / f"{req_base}{config.FINAL_WORD_SUFFIX}" # 重新获取路径确保准确
            if word_to_move.exists():
                try:
                    shutil.move(str(word_to_move), str(final_result_dir / word_to_move.name))
                    logger.info(f"[{req_name}] 已移动文件: {word_to_move.name} 到 {final_result_dir.name}")
                except Exception as move_word_err:
                    logger.error(f"[{req_name}] 移动 Word 文件 {word_to_move.name} 失败: {move_word_err}", exc_info=True)
            else:
                 # 检查 final_word_output 是否存在于 locals() 以防万一
                 if 'final_word_output' in locals() and final_word_output.exists():
                     try:
                         shutil.move(str(final_word_output), str(final_result_dir / final_word_output.name))
                         logger.info(f"[{req_name}] 已移动文件 (备用路径): {final_word_output.name} 到 {final_result_dir.name}")
                     except Exception as move_word_err_alt:
                         logger.error(f"[{req_name}] 移动 Word 文件 (备用路径) {final_word_output.name} 失败: {move_word_err_alt}", exc_info=True)
                 else:
                     logger.warning(f"[{req_name}] 未找到最终 Word 文件进行移动: {word_to_move.name}")


        except Exception as final_org_err:
            logger.error(f"[{req_name}] 创建 final_result 目录或移动文件时出错: {final_org_err}", exc_info=True)


    except Exception as post_process_e: # This corresponds to the main try block
        logger.error(f"[{req_name}] 阶段 4-8 后置处理失败: {post_process_e}", exc_info=True) # 更新日志消息中的阶段号
        # Decide if this should halt the entire process for this requirement
        # raise post_process_e # Optionally re-raise
